# -*- coding: utf-8 -*-
"""
Excel导出模块：将匹配结果写入多工作表的Excel文件。
支持 1-10 个来源的交集统计导出。
"""
import logging
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 颜色配置
_HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F7')
_SUBHEADER_FONT = Font(bold=True, color='1F3864', size=10)
_BORDER_SIDE = Side(style='thin', color='AAAAAA')
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE
)
_ALT_FILL = PatternFill('solid', fgColor='F2F7FB')


def _style_header(cell):
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _CELL_BORDER


def _style_subheader(cell):
    cell.fill = _SUBHEADER_FILL
    cell.font = _SUBHEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _CELL_BORDER


def _style_data(cell, alt=False):
    if alt:
        cell.fill = _ALT_FILL
    cell.alignment = Alignment(vertical='center')
    cell.border = _CELL_BORDER


def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _autosize_columns(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(row_idx, col_idx).value or '')) for row_idx in range(1, ws.max_row + 1)),
            default=10
        )
        _set_col_width(ws, col_idx, min(max_len + 4, max_width))


def _write_sheet(ws, rows: List[Dict], columns):
    ws.freeze_panes = 'A2'
    for col_idx, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header(cell)

    for row_idx, row in enumerate(rows, 2):
        alt = row_idx % 2 == 0
        for col_idx, (_, accessor) in enumerate(columns, 1):
            value = accessor(row) if callable(accessor) else row.get(accessor, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style_data(cell, alt)

    _autosize_columns(ws)


def _write_summary_sheet(ws, result: Dict):
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 18
    row = 1

    def write_header(text):
        nonlocal row
        cell = ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _style_header(cell)
        row += 1

    def write_row(label, value, alt=False):
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        _style_data(c1, alt)
        _style_data(c2, alt)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    write_header('各数据库有效期刊数')
    for i, (src, cnt) in enumerate(result['counts'].items()):
        write_row('  ' + src, cnt, i % 2 == 0)

    row += 1
    write_header('交集统计')

    intersections = result.get('intersections', {})
    idx = 0
    for size in sorted(intersections.keys(), reverse=True):
        for combo_key, entries in sorted(intersections[size].items()):
            write_row(combo_key + ' 交集', len(entries), idx % 2 == 0)
            idx += 1

    row += 1
    write_header('单库独有统计')
    for i, (src, entries) in enumerate(result.get('one_only', {}).items()):
        write_row('仅在' + src + '中收录', len(entries), i % 2 == 0)


def _flatten_entries(grouped_entries: Dict, group_field: str) -> List[Dict]:
    rows = []
    for group_name, entries in grouped_entries.items():
        for entry in entries:
            row = dict(entry)
            row[group_field] = group_name
            rows.append(row)
    return rows


def _journal_columns(group_label='来源组合'):
    return [
        ('序号', lambda row: row.get('_index', '')),
        ('刊名', 'raw_name'),
        ('标准名', 'normalized_name'),
        ('匹配键', 'key'),
        (group_label, 'group_name'),
        ('来源库', lambda row: ' + '.join(row.get('sources', []))),
    ]


def _prepare_rows(entries: List[Dict], group_name='') -> List[Dict]:
    rows = []
    for index, entry in enumerate(entries, 1):
        row = dict(entry)
        row['_index'] = index
        if group_name:
            row['group_name'] = group_name
        else:
            row['group_name'] = row.get('group_name', '')
        rows.append(row)
    return rows


def export(result: Dict, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet('统计摘要')
    _write_summary_sheet(ws_summary, result)
    logger.info('已写入工作表：统计摘要')

    intersections = result.get('intersections', {})
    for size in sorted(intersections.keys(), reverse=True):
        sheet_name = str(size) + '库交集'
        all_entries = []
        for combo_key, entries in sorted(intersections[size].items()):
            for entry in entries:
                row = dict(entry)
                row['group_name'] = combo_key
                all_entries.append(row)
        
        if all_entries or size == len(result.get('db_names', [])):
            ws = wb.create_sheet(sheet_name)
            _write_sheet(ws, _prepare_rows(all_entries), _journal_columns('交集组合'))
            logger.info('已写入工作表：' + sheet_name + '（' + str(len(all_entries)) + ' 种）')

    one_rows = _flatten_entries(result.get('one_only', {}), 'group_name')
    ws = wb.create_sheet('单库独有')
    _write_sheet(ws, _prepare_rows(one_rows), _journal_columns('独有来源'))
    logger.info('已写入工作表：单库独有（' + str(len(one_rows)) + ' 条）')

    wb.save(output_path)
    logger.info('Excel 已保存：' + output_path)
