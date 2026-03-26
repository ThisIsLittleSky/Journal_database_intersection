# -*- coding: utf-8 -*-
"""
北大核心期刊解析器。
文件格式：单个工作表，第1行总标题，第2行列头，第3行起数据。
刊名在第D列（index 3，0-based）。
"""
import logging
from openpyxl import load_workbook
from core.normalizer import normalize

logger = logging.getLogger(__name__)


def parse(filepath: str) -> dict:
    """
    返回 {'journals': [{'name': 原始刊名, 'key': 标准化刊名, 'source': '北大核心'}]}
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    results = []
    seen_keys = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)

        # 扫描找到列头行（含"刊名"的行）
        header_row_idx = None
        name_col_idx = None
        peeked = []

        for i, row in enumerate(rows):
            peeked.append(row)
            if row is None:
                continue
            for j, cell in enumerate(row):
                if cell and '刊名' in str(cell):
                    header_row_idx = i
                    name_col_idx = j
                    break
            if header_row_idx is not None:
                break

        if name_col_idx is None:
            # 未找到列头，默认第2行(index 1)是列头，刊名在index 3
            name_col_idx = 3
            data_start = 2
            peeked_iter = iter(peeked)
        else:
            data_start = header_row_idx + 1
            peeked_iter = iter(peeked[data_start:])

        # 读数据行
        for row in peeked_iter:
            if row is None or len(row) <= name_col_idx:
                continue
            name = row[name_col_idx]
            if not name or not isinstance(name, str):
                continue
            name = name.strip()
            if not name:
                continue
            key = normalize(name)
            if not key or key in seen_keys:
                continue
            seen_keys.add(key)
            results.append({'name': name, 'key': key, 'source': '北大核心'})

        # 继续读剩余行（peeked_iter 已消费 peeked 之后的行，需继续读 ws）
        for row in rows:
            if row is None or len(row) <= name_col_idx:
                continue
            name = row[name_col_idx]
            if not name or not isinstance(name, str):
                continue
            name = name.strip()
            if not name:
                continue
            key = normalize(name)
            if not key or key in seen_keys:
                continue
            seen_keys.add(key)
            results.append({'name': name, 'key': key, 'source': '北大核心'})

    wb.close()
    logger.info(f'北大核心：解析完成，有效期刊数 {len(results)}')
    return {'journals': results}
