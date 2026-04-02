# -*- coding: utf-8 -*-
import logging
import os

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_HEADER_KEYWORDS = ('刊名', '期刊', 'journal', 'title', 'source title')
_EXCEL_CHUNK_ROWS = 500
_MAX_STRUCTURED_FIELDS = 6
_MAX_CELL_TEXT = 160
_TABULAR_INPUT_VERSION = 'tabular_candidates_v1'


def _looks_like_name(value) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text or len(text) < 2:
        return False
    if text.isdigit():
        return False
    return True


def _detect_name_column(rows: list[tuple]) -> tuple[int, int]:
    for row_idx, row in enumerate(rows[:10]):
        for col_idx, cell in enumerate(row):
            text = str(cell or '').strip().lower()
            if any(keyword in text for keyword in _HEADER_KEYWORDS):
                return col_idx, row_idx + 1

    scores = {}
    for row in rows[:30]:
        for col_idx, cell in enumerate(row):
            if _looks_like_name(cell):
                scores[col_idx] = scores.get(col_idx, 0) + 1

    if not scores:
        return 0, 0
    best_col = max(scores.items(), key=lambda item: item[1])[0]
    return best_col, 0


def _build_preview(rows: list[tuple], max_rows: int = 60) -> str:
    preview_lines = []
    for row in rows[:max_rows]:
        cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if cells:
            preview_lines.append('\t'.join(cells))
    return '\n'.join(preview_lines)


def _build_sheet_lines(rows: list[tuple]) -> list[str]:
    lines = []
    for row in rows:
        cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if cells:
            lines.append('\t'.join(cells))
    return lines


def _build_sheet_chunks(sheet_name: str, rows: list[tuple], chunk_size: int = _EXCEL_CHUNK_ROWS) -> list[dict]:
    lines = _build_sheet_lines(rows)
    chunks = []
    for start in range(0, len(lines), chunk_size):
        batch = lines[start:start + chunk_size]
        if not batch:
            continue
        start_row = start + 1
        end_row = start + len(batch)
        chunks.append({
            'label': f'[{sheet_name}] 第{start_row}-{end_row}行',
            'text': '\n'.join(batch),
        })
    return chunks


def _get_header_name(headers: list[str], col_idx: int) -> str:
    if col_idx < len(headers) and headers[col_idx]:
        return headers[col_idx]
    return f'col_{col_idx + 1}'


def _build_headers(rows: list[tuple], data_start: int) -> list[str]:
    max_cols = max((len(row) for row in rows[:30]), default=0)
    headers = [f'col_{idx + 1}' for idx in range(max_cols)]
    if data_start <= 0 or data_start > len(rows):
        return headers

    header_row = rows[data_start - 1]
    return [
        str(header_row[idx]).strip() if idx < len(header_row) and str(header_row[idx]).strip() else headers[idx]
        for idx in range(max_cols)
    ]


def _clean_cell_text(value, max_length: int = _MAX_CELL_TEXT) -> str:
    text = str(value or '').strip()
    if not text:
        return ''
    text = ' '.join(text.split())
    return text[:max_length]


def _build_structured_cells(headers: list[str], row: tuple, name_col_idx: int) -> dict:
    pairs = []

    if len(row) > name_col_idx:
        candidate_text = _clean_cell_text(row[name_col_idx])
        if candidate_text:
            pairs.append((_get_header_name(headers, name_col_idx), candidate_text))

    for col_idx, value in enumerate(row):
        if col_idx == name_col_idx:
            continue
        text = _clean_cell_text(value)
        if not text:
            continue
        pairs.append((_get_header_name(headers, col_idx), text))
        if len(pairs) >= _MAX_STRUCTURED_FIELDS:
            break

    return {key: value for key, value in pairs[:_MAX_STRUCTURED_FIELDS]}


def parse(filepath: str) -> dict:
    source_name = os.path.splitext(os.path.basename(filepath))[0]
    wb = load_workbook(filepath, read_only=True, data_only=True)
    results = []
    seen_names = set()
    preview_chunks = []
    raw_chunks = []
    structured_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        preview = _build_preview(rows)
        if preview:
            preview_chunks.append(f'[{sheet_name}]\n{preview}')
        raw_chunks.extend(_build_sheet_chunks(sheet_name, rows))

        name_col_idx, data_start = _detect_name_column(rows)
        headers = _build_headers(rows, data_start)
        for row_index, row in enumerate(rows[data_start:], start=data_start + 1):
            if row is None or len(row) <= name_col_idx:
                continue
            value = row[name_col_idx]
            if not _looks_like_name(value):
                continue
            name = str(value).strip()
            if not name or name in seen_names:
                continue
            seen_names.add(name)
            results.append({'name': name, 'key': name, 'source': source_name})
            structured_rows.append({
                'sheet': sheet_name,
                'row_index': row_index,
                'candidate_name': name,
                'cells': _build_structured_cells(headers, row, name_col_idx),
            })

    wb.close()
    logger.info(f'通用Excel解析完成：{source_name}，有效期刊数 {len(results)}')
    return {
        'source_db': source_name,
        'journals': results,
        'raw_text': '\n\n'.join(preview_chunks)[:12000],
        'raw_chunks': raw_chunks,
        'raw_chunk_policy': 'excel_rows_500_v1',
        'structured_rows': structured_rows,
        'structured_input_type': 'tabular',
        'structured_input_version': _TABULAR_INPUT_VERSION,
    }
