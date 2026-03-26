# -*- coding: utf-8 -*-
"""
CSSCI解析器。
文件格式：多个工作表，每表：
  第1行：'CSSCI来XX种' | ... | 'CSSCI扩展版（X种）' ...
  第2行：序号 | 刊名（主刊） | 空 | 序号 | 刊名（扩展版）
  第3行起：数据
只取主刊列（B列，index 1），完全排除扩展版列。
"""
import logging
from openpyxl import load_workbook
from core.normalizer import normalize

logger = logging.getLogger(__name__)


def _find_main_col(header_row) -> int:
    """在第2行中找主刊刊名列索引（第一个含'刊名'的列）。"""
    if not header_row:
        return 1
    for j, cell in enumerate(header_row):
        if cell and '刊名' in str(cell):
            return j
    return 1


def _find_ext_start_col(first_row) -> int:
    """在第1行中找扩展版起始列索引（含'扩展'的列），返回 None 表示无扩展版。"""
    if not first_row:
        return None
    for j, cell in enumerate(first_row):
        if cell and '扩展' in str(cell):
            return j
    return None


def parse(filepath: str) -> dict:
    """
    返回 {'journals': [{'name': 原始刊名, 'key': 标准化刊名, 'source': 'CSSCI'}]}
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    results = []
    seen_keys = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 3:
            continue

        first_row = all_rows[0]   # 第1行：种数说明
        header_row = all_rows[1]  # 第2行：列头

        main_col = _find_main_col(header_row)
        ext_start_col = _find_ext_start_col(first_row)

        for row in all_rows[2:]:
            if row is None or len(row) <= main_col:
                continue
            # 只取主刊列，且该列不能是扩展版列
            if ext_start_col is not None and main_col >= ext_start_col:
                # 列定位出错，退回默认
                main_col = 1

            name = row[main_col]
            if not name or not isinstance(name, str):
                continue
            name = name.strip()
            if not name:
                continue
            key = normalize(name)
            if not key or key in seen_keys:
                continue
            seen_keys.add(key)
            results.append({'name': name, 'key': key, 'source': 'CSSCI'})

    wb.close()
    logger.info(f'CSSCI：解析完成，有效期刊数（主刊，已排除扩展版）{len(results)}')
    return {'journals': results}
